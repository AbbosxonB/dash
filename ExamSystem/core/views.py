import pandas as pd # For Excel reading
import csv
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponse
from .models import CustomUser, Test, Subject, StudentResult, Question, Answer
from .decorators import role_required, redirect_based_on_role
import openpyxl # type: ignore
import csv
from django.utils import timezone


@role_required(['Teacher'])
def import_test_view(request):
    """
    View for teachers to import tests from an Excel file.
    """
    if request.method == 'POST':
        test_name = request.POST.get('test_name')
        subject_id = request.POST.get('subject')
        total_time = request.POST.get('total_time_minutes')
        default_points_value = request.POST.get('default_points_value')
        excel_file = request.FILES.get('excel_file')

        if not all([test_name, subject_id, total_time, default_points_value, excel_file]):
            messages.error(request, 'All fields are required, including the Excel file and Points per Question.')
            return redirect('import_test')

        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            messages.error(request, 'Selected subject does not exist.')
            return redirect('import_test')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Invalid file type. Please upload an Excel file (.xls or .xlsx).')
            return redirect('import_test')

        try:
            # Create the Test object first
            test = Test.objects.create(
                test_name=test_name,
                subject=subject,
                created_by=request.user,
                total_time_minutes=int(total_time),
                status='Draft' # Imported tests are initially in Draft
            )

            df = pd.read_excel(excel_file)

            # Expected columns: 'Question Text', 'Question Type', 'Answer 1 Text', 'Answer 1 Correct', ...
            # I will assume a structure where answers are in columns like 'Answer X Text' and 'Answer X Correct'
            
            for index, row in df.iterrows():
                question_text = row.get("Question Text (Only 'MCQ' supported)")
                question_type = row.get('Question Type', 'MCQ').upper() # Default to MCQ, but will enforce MCQ below

                if not question_text:
                    messages.warning(request, f'Skipping row {index + 2} due to missing question text.')
                    continue
                
                # Ensure only MCQ questions are imported (even if Excel specified otherwise)
                question = Question.objects.create(
                    test=test,
                    question_text=question_text,
                    question_type='MCQ', # Always 'MCQ' as per requirement
                    points_value=int(default_points_value) # Use the default points value from the form
                )

                # Process answers (assuming up to 6 answers per question)
                for i in range(1, 7): # Check for Answer 1 to Answer 6
                    answer_text = row.get(f'Answer {i} Text')
                    is_correct = row.get(f'Answer {i} Correct', False) # Default to False

                    if answer_text:
                        Answer.objects.create(
                            question=question,
                            answer_text=answer_text,
                            is_correct=bool(is_correct) # Ensure boolean
                        )
            
            messages.success(request, f'Test "{test_name}" imported successfully with {df.shape[0]} questions!')
            return redirect('teacher_dashboard')

        except Exception as e:
            messages.error(request, f'Error importing test: {e}')
            # If test was created but then an error occurred with questions, delete the test
            if 'test' in locals() and test.questions.count() == 0:
                test.delete()
            return redirect('import_test')

    # For GET request, display the form
    subjects = Subject.objects.filter(created_by=request.user)
    context = {
        'subjects': subjects,
        'user_role': request.user.role
    }
    return render(request, 'core/import_test.html', context)


def login_view(request):
    """
    Login view that handles authentication and redirects based on user role.
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Redirect based on user role
            if user.role == 'Admin':
                return redirect('admindashboard')  # This will be defined in URLs
            elif user.role == 'Teacher':
                return redirect('teacher_dashboard')
            elif user.role == 'Student':
                return redirect('student_dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'core/login.html')


@login_required
def logout_view(request):
    """
    Logout view
    """
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    """
    Generic dashboard view that redirects based on user role
    """
    user_role = request.user.role
    
    if user_role == 'Admin':
        return redirect('admindashboard')
    elif user_role == 'Teacher':
        return redirect('teacher_dashboard')
    elif user_role == 'Student':
        return redirect('student_dashboard')
    else:
        return redirect('login')


@role_required(['Admin', 'Teacher'])
def admin_dashboard_view(request):
    """
    Admin dashboard view - accessible by Admin and Teacher (for admin-like functions)
    """
    # Get statistics for the dashboard
    total_users = CustomUser.objects.count()
    total_tests = Test.objects.count()
    total_subjects = Subject.objects.count()
    total_results = StudentResult.objects.count()
    
    # Get all student results
    student_results = StudentResult.objects.all()
    
    context = {
        'total_users': total_users,
        'total_tests': total_tests,
        'total_subjects': total_subjects,
        'total_results': total_results,
        'student_results': student_results,
        'user_role': request.user.role
    }
    return render(request, 'core/admin_dashboard.html', context)


@role_required(['Admin', 'Teacher'])
def export_student_results_xls(request):
    """
    Export all student results to an Excel file.
    """
    results = StudentResult.objects.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"
    
    # Add header row
    headers = ["Student", "Test", "Score Achieved", "Total Score", "Time Taken (s)", "Completion Date"]
    ws.append(headers)
    
    # Add data rows
    for result in results:
        ws.append([
            result.student.username,
            result.test.test_name,
            result.score_achieved,
            result.total_score,
            result.time_taken,
            result.completion_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=student_results.xlsx'
    wb.save(response)
    
    return response


@role_required(['Admin', 'Teacher'])
def export_student_results_csv(request):
    """
    Export all student results to a CSV file.
    """
    results = StudentResult.objects.all()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=student_results.csv'
    
    writer = csv.writer(response)
    # Add header row
    writer.writerow(["Student", "Test", "Score Achieved", "Total Score", "Time Taken (s)", "Completion Date"])
    
    # Add data rows
    for result in results:
        writer.writerow([
            result.student.username,
            result.test.test_name,
            result.score_achieved,
            result.total_score,
            result.time_taken,
            result.completion_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
        
    return response


@role_required(['Teacher'])
def teacher_dashboard_view(request):
    """
    Teacher dashboard view
    """
    # Get tests created by this teacher
    teacher_tests = Test.objects.filter(created_by=request.user)
    
    # Get subjects created by this teacher
    teacher_subjects = Subject.objects.filter(created_by=request.user)
    
    context = {
        'tests': teacher_tests,
        'subjects': teacher_subjects,
        'user_role': request.user.role
    }
    return render(request, 'core/teacher_dashboard.html', context)


@role_required(['Student'])
def student_dashboard_view(request):
    """
    Student dashboard view
    """
    # Get available subjects and tests
    available_tests = Test.objects.filter(status='Published')
    
    # Get student's results
    student_results = StudentResult.objects.filter(student=request.user)
    
    context = {
        'available_tests': available_tests,
        'student_results': student_results,
        'user_role': request.user.role
    }
    return render(request, 'core/student_dashboard.html', context)


@role_required(['Teacher'])
def create_test_view(request):
    """
    View for creating a test with questions and answers
    """
    if request.method == 'POST':
        # Process form data to create test with questions and answers
        test_name = request.POST.get('test_name')
        subject_id = request.POST.get('subject')
        total_time = request.POST.get('total_time_minutes')
        status = request.POST.get('status', 'Draft')
        
        # Get subject object
        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            messages.error(request, 'Subject does not exist.')
            return redirect('create_test')
        
        # Get the default points value for all questions
        default_points_value = int(request.POST.get('default_points_value', 1))

        # Create the test
        test = Test.objects.create(
            test_name=test_name,
            subject=subject,
            created_by=request.user,
            total_time_minutes=int(total_time),
            status=status
        )
        
        # Process questions and answers
        question_texts = request.POST.getlist('question_text')
        
        for i, question_text in enumerate(question_texts):
            if question_text.strip():  # Only create if question text is not empty
                from .models import Question, Answer # DIAGNOSTIC IMPORT
                
                question = Question.objects.create(
                    test=test,
                    question_text=question_text,
                    question_type='MCQ', # Always MCQ as per requirement
                    points_value=default_points_value # Use the default points value
                )
                
                # Process answers for this question
                answer_texts = request.POST.getlist(f'question_{i}_answer_text')
                is_correct_list = request.POST.getlist(f'question_{i}_is_correct')
                
                for j, answer_text in enumerate(answer_texts):
                    if answer_text.strip():  # Only create if answer text is not empty
                        is_correct = str(j) in is_correct_list  # Check if this answer is marked as correct
                        Answer.objects.create(
                            question=question,
                            answer_text=answer_text,
                            is_correct=is_correct
                        )
        
        messages.success(request, 'Test created successfully!')
        return redirect('teacher_dashboard')
    
    # For GET request, show the form
    subjects = Subject.objects.filter(created_by=request.user)  # Only subjects created by the teacher
    context = {
        'subjects': subjects,
        'user_role': request.user.role
    }
    return render(request, 'core/create_test.html', context)


@role_required(['Student'])
def take_test_view(request, test_id):
    """
    View for students to take a test
    """
    try:
        test = Test.objects.get(id=test_id, status='Published')  # Only published tests
    except Test.DoesNotExist:
        messages.error(request, 'Test does not exist or is not available.')
        return redirect('student_dashboard')

    # Check if the student has already completed this test or has a pending retake
    pending_results = StudentResult.objects.filter(student=request.user, test=test, status='Pending')
    completed_results = StudentResult.objects.filter(student=request.user, test=test, status='Completed')

    if completed_results.exists() and not pending_results.exists():
        messages.error(request, f"You have already completed the test for {test.subject.name} and no retake attempt is currently available.")
        return redirect('student_dashboard')
    
    # If a pending result exists, or if no results exist (first attempt), proceed.
    
    # Get questions for the test
    from django.db.models import Q
    import random
    all_questions = list(test.questions.all())
    
    # Randomly select up to 25 questions
    if len(all_questions) > 25:
        questions = random.sample(all_questions, 25)
    else:
        questions = all_questions
    
    random.shuffle(questions)  # Randomize the selected questions
    
    context = {
        'test': test,
        'questions': questions,
        'user_role': request.user.role
    }
    return render(request, 'core/take_test.html', context)


@role_required(['Student'])
def submit_test_view(request, test_id):
    """
    View to handle test submission and scoring
    """
    if request.method != 'POST':
        return redirect('student_dashboard')
    
    try:
        test = Test.objects.get(id=test_id, status='Published')
    except Test.DoesNotExist:
        messages.error(request, 'Test does not exist or is not available.')
        return redirect('student_dashboard')
    
    # Calculate score
    score = 0
    total_score = 0
    
    for question in test.questions.all():
        total_score += question.points_value
        selected_answer_id = request.POST.get(f'question_{question.id}')
        
        if selected_answer_id:
            try:
                selected_answer = Answer.objects.get(id=selected_answer_id, question=question)
                if selected_answer.is_correct:
                    score += question.points_value
            except Answer.DoesNotExist:
                pass  # Invalid answer ID, skip
    
    # Save or update the result
    from django.utils import timezone
    completion_time = request.POST.get('time_taken', 0)  # Time in seconds
    
    # Try to find an existing pending result for this student and test
    student_result = StudentResult.objects.filter(
        student=request.user, 
        test=test, 
        status='Pending'
    ).first()
    total_score = 25
    if student_result:
        # Update the existing pending result
        student_result.score_achieved = score
        student_result.total_score = total_score
        student_result.time_taken = int(completion_time)
        student_result.completion_date = timezone.now()
        student_result.status = 'Completed'
        student_result.save()
    else:
        # Create a new result if no pending one exists (e.g., first attempt)
        StudentResult.objects.create(
            student=request.user,
            test=test,
            score_achieved=score,
            total_score=total_score,
            time_taken=int(completion_time),
            completion_date=timezone.now(),
            status='Completed'
        )
    
    messages.success(request, f'Test submitted successfully! Your score: {score}/{total_score}')
    return redirect('student_dashboard')


@role_required(['Teacher'])
def edit_test_view(request, test_id):
    """
    View for editing an existing test.
    """
    try:
        test = Test.objects.get(id=test_id, created_by=request.user)
    except Test.DoesNotExist:
        messages.error(request, 'Test not found or you do not have permission to edit it.')
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        # Update test details
        test.test_name = request.POST.get('test_name')
        subject_id = request.POST.get('subject')
        try:
            test.total_time_minutes = int(request.POST.get('total_time_minutes', 0))
        except ValueError:
            messages.error(request, 'Invalid value for total time. Please enter a number.')
            return redirect('edit_test', test_id=test.id)
        test.status = request.POST.get('status')
        
        try:
            subject = Subject.objects.get(id=subject_id)
            test.subject = subject
        except Subject.DoesNotExist:
            messages.error(request, 'Invalid subject selected.')
            return redirect('edit_test', test_id=test.id)

        test.save()

        # Get the default points value for all questions
        default_points_value = int(request.POST.get('default_points_value', 1))

        # Update questions and answers
        # Safely convert question_ids to integers, ignoring any non-integer or empty strings
        parsed_question_ids = []
        for q_id_str in request.POST.getlist('question_id'):
            if q_id_str: # Ensure it's not an empty string before attempting conversion
                try:
                    parsed_question_ids.append(int(q_id_str))
                except ValueError:
                    messages.warning(request, f"Invalid question ID encountered and skipped: '{q_id_str}'.")
        
        question_texts = request.POST.getlist('question_text')
        
        # Delete questions that are not in the form
        # Ensure that parsed_question_ids is not empty to avoid issues with exclude if no valid IDs are present
        if parsed_question_ids:
            test.questions.exclude(id__in=parsed_question_ids).delete()
        else:
            # If no question IDs are submitted, it means all existing questions should be deleted
            # This handles cases where all questions might be removed from the frontend
            if not question_texts: # Only delete all if no new question texts are provided either
                test.questions.all().delete()

        # Before iterating, ensure question_ids used for indexing matches the parsed_question_ids length
        # This part of the code needs to be careful because question_ids and question_texts are parallel lists
        # We need to map `question_texts` to `parsed_question_ids` based on their original order from POST
        # If a question_id was invalid, it was skipped, so the index in question_texts won't directly map.
        # A more robust approach might be to associate question_id with question_text in the loop.
        
        # Let's re-think the loop for updating/creating questions.
        # We need to iterate through the submitted question_texts and try to match them with existing questions
        # or create new ones.
        
        existing_question_ids = set(test.questions.values_list('id', flat=True))
        
        for i, question_text in enumerate(question_texts):
            # Check if there's a corresponding question_id from the form data
            # Use request.POST.getlist('question_id') directly here to match indices with question_texts
            # and then validate/parse the ID within the loop.
            submitted_question_id_str = request.POST.getlist('question_id')[i] if i < len(request.POST.getlist('question_id')) else ''
            
            question_obj = None
            if submitted_question_id_str:
                try:
                    submitted_question_id = int(submitted_question_id_str)
                    if submitted_question_id in existing_question_ids:
                        question_obj = test.questions.get(id=submitted_question_id)
                except (ValueError, Question.DoesNotExist):
                    messages.warning(request, f"Submitted question ID '{submitted_question_id_str}' was invalid or did not exist. Treating as new question.")
            
            if question_text.strip(): # Only process if question text is not empty
                if question_obj:
                    # Update existing question
                    question_obj.question_text = question_text
                    question_obj.question_type = 'MCQ' # Enforce MCQ
                    question_obj.points_value = default_points_value # Use the default points value
                    question_obj.save()
                else:
                    # Create new question
                    question_obj = Question.objects.create(
                        test=test,
                        question_text=question_text,
                        question_type='MCQ', # Enforce MCQ
                        points_value=default_points_value # Use the default points value
                    )

                # Update answers for the question
                # This part assumes answers are submitted per question based on their index.
                # It's important that the frontend sends answer data correctly aligned with question data.
                answer_ids_for_question = request.POST.getlist(f'question_{i}_answer_id')
                answer_texts_for_question = request.POST.getlist(f'question_{i}_answer_text')
                correct_answers_for_question = request.POST.getlist(f'question_{i}_is_correct')
                
                # Filter out invalid answer IDs for this specific question
                parsed_answer_ids = []
                for ans_id_str in answer_ids_for_question:
                    if ans_id_str:
                        try:
                            parsed_answer_ids.append(int(ans_id_str))
                        except ValueError:
                            messages.warning(request, f"Invalid answer ID encountered and skipped: '{ans_id_str}'.")

                # Delete answers that are not in the form for this question
                if parsed_answer_ids:
                    question_obj.answers.exclude(id__in=parsed_answer_ids).delete()
                else:
                    # If no answer IDs are submitted for this question, delete all existing answers for it
                    if not answer_texts_for_question:
                        question_obj.answers.all().delete()

                for j, answer_text in enumerate(answer_texts_for_question):
                    submitted_answer_id_str = answer_ids_for_question[j] if j < len(answer_ids_for_question) else ''
                    is_correct = str(j) in correct_answers_for_question

                    answer_obj = None
                    if submitted_answer_id_str:
                        try:
                            submitted_answer_id = int(submitted_answer_id_str)
                            answer_obj = question_obj.answers.get(id=submitted_answer_id)
                        except (ValueError, Answer.DoesNotExist):
                            messages.warning(request, f"Submitted answer ID '{submitted_answer_id_str}' was invalid or did not exist. Treating as new answer.")

                    if answer_text.strip(): # Only process if answer text is not empty
                        if answer_obj:
                            # Update existing answer
                            answer_obj.answer_text = answer_text
                            answer_obj.is_correct = is_correct
                            answer_obj.save()
                        else:
                            # Create new answer
                            Answer.objects.create(
                                question=question_obj,
                                answer_text=answer_text,
                                is_correct=is_correct
                            )
            else:
                # If question_text is empty, and it was an existing question, ensure it's deleted.
                # This case should ideally be handled by the initial test.questions.exclude logic
                # but adding a safeguard here for a submitted_question_id that might not have a text.
                if question_obj and question_obj.id in existing_question_ids:
                    question_obj.delete()



        messages.success(request, 'Test updated successfully!')
        return redirect('teacher_dashboard')

    # GET request
    subjects = Subject.objects.filter(created_by=request.user)
    context = {
        'test': test,
        'subjects': subjects,
        'user_role': request.user.role
    }
    return render(request, 'core/edit_test.html', context)


@role_required(['Teacher'])
def delete_test_view(request, test_id):
    """
    View to delete a test.
    """
    try:
        test = Test.objects.get(id=test_id, created_by=request.user)
    except Test.DoesNotExist:
        messages.error(request, 'Test not found or you do not have permission to delete it.')
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        test.delete()
        messages.success(request, 'Test deleted successfully!')
        return redirect('teacher_dashboard')
    
    context = {
        'test': test,
        'user_role': request.user.role,
        'object_type': 'Test'
    }
    return render(request, 'core/confirm_delete.html', context)


@role_required(['Admin'])
def delete_student_result_view(request, result_id):
    """
    View to delete a student result, allowing them to retake the test.
    """
    try:
        result = StudentResult.objects.get(id=result_id)
    except StudentResult.DoesNotExist:
        messages.error(request, 'Student result not found.')
        return redirect('admindashboard')

    if request.method == 'POST':
        result.delete()
        messages.success(request, f"Result for {result.student.username} on test '{result.test.test_name}' has been deleted. The student can now retake the test.")
        return redirect('admindashboard')
    
    context = {
        'object': result,
        'user_role': request.user.role,
        'object_type': 'Student Result'
    }
    return render(request, 'core/confirm_delete.html', context)


@role_required(['Admin'])
def retake_test_view(request, result_id):
    """
    View to initiate a test retake by creating a new pending student result.
    The previous result is preserved.
    """
    try:
        original_result = StudentResult.objects.get(id=result_id)
    except StudentResult.DoesNotExist:
        messages.error(request, 'Original student result not found.')
        return redirect('admindashboard')

    # Create a new StudentResult instance for the retake
    # The new result will have default score, time_taken, and completion_date as null
    # and its status will be 'Pending' by default.
    new_result = StudentResult.objects.create(
        student=original_result.student,
        test=original_result.test,
        # score_achieved, total_score, time_taken, completion_date are null by default
        # status is 'Pending' by default
    )
    
    messages.success(request, f"A new retake attempt for {original_result.student.username} on test '{original_result.test.test_name}' has been initiated. The student can now take the test again.")
    return redirect('admindashboard')


@role_required(['Teacher'])
def download_sample_excel(request):
    """
    Generates and serves a sample Excel file for test import.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions and Answers"

    ws.append([
        "Question Text (Only 'MCQ' supported)", "Answer 1 Text", "Answer 1 Correct",
        "Answer 2 Text", "Answer 2 Correct",
        "Answer 3 Text", "Answer 3 Correct",
        "Answer 4 Text", "Answer 4 Correct",
        "Answer 5 Text", "Answer 5 Correct",
        "Answer 6 Text", "Answer 6 Correct",
    ])

    # Add example data (only MCQ)
    ws.append([
        "What is the capital of France?", "Paris", True,
        "London", False,
        "Berlin", False,
        "Rome", False,
        "", "", # Empty answers
        "", ""
    ])


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=sample_test_import.xlsx'
    wb.save(response)
    return response


@role_required(['Admin'])
def import_students_view(request):
    """
    View for admins to import students from an Excel file.
    """
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, 'Excel file is required.')
            return redirect('import_students')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Invalid file type. Please upload an Excel file (.xls or .xlsx).')
            return redirect('import_students')

        try:
            df = pd.read_excel(excel_file)
            
            required_columns = ['student_id', 'full_name', 'passport_series', 'course', 'group', 'direction']
            if not all(col in df.columns for col in required_columns):
                messages.error(request, f'Excel file must contain the following columns: {", ".join(required_columns)}')
                return redirect('import_students')

            for index, row in df.iterrows():
                student_id = row['student_id']
                full_name = row['full_name']
                passport_series = row['passport_series']
                course = row['course']
                group = row['group']
                direction = row['direction']

                if CustomUser.objects.filter(username=student_id).exists():
                    messages.warning(request, f'Student with ID {student_id} already exists. Skipping.')
                    continue

                user = CustomUser.objects.create_user(
                    username=student_id,
                    password=str(passport_series),
                    role='Student',
                    student_id=student_id,
                    student_full_name=full_name,
                    course=course,
                    student_groups=group,
                    student_direction=direction
                )
                user.save()

            messages.success(request, f'Successfully imported {len(df)} students.')
            return redirect('admindashboard')

        except Exception as e:
            messages.error(request, f'Error importing students: {e}')
            return redirect('import_students')

    return render(request, 'core/import_students.html', {'user_role': request.user.role})


@role_required(['Admin'])
def download_sample_student_excel(request):
    """
    Generates and serves a sample Excel file for student import.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append([
        'student_id', 'full_name', 'passport_series', 'course', 'group', 'direction'
    ])

    ws.append([
        '123456', 'John Doe', 'AB1234567', '2', 'A', 'Computer Science'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=sample_student_import.xlsx'
    wb.save(response)
    return response

